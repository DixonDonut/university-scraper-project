# scraper/excel_writer.py
# Builds Excel output per PROJECT_RULES.md Section 6

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Colour fills
FILLS = {
    'header':    PatternFill("solid", fgColor="1F3864"),
    'nmc_yes':   PatternFill("solid", fgColor="FFF2CC"),
    'diploma':   PatternFill("solid", fgColor="FCE4D6"),
    'english1':  PatternFill("solid", fgColor="E2EFDA"),
    'english2':  PatternFill("solid", fgColor="EBF5E1"),
    'bachelor1': PatternFill("solid", fgColor="D6E4F0"),
    'bachelor2': PatternFill("solid", fgColor="FFFFFF"),
}

# ── Fonts
FONTS = {
    'header':   Font(color="FFFFFF", bold=True, size=11),
    'nmc':      Font(color="7F6000", size=10),
    'english':  Font(color="375623", size=10),
    'diploma':  Font(color="833C00", size=10),
    'url':      Font(color="0563C1", underline="single", size=10),
    'standard': Font(color="000000", size=10),
}

HEADERS = [
    "Course Name",
    "Degree Level",
    "Programme Page URL",
    "Duration",
    "Tuition Fee",
    "Medium of Instruction",
    "Needs Manual Check",
]

COLUMN_WIDTHS = {
    'A': 55,   # Course Name
    'B': 14,   # Degree Level
    'C': 65,   # URL
    'D': 12,   # Duration
    'E': 22,   # Fee
    'F': 22,   # Medium
    'G': 20,   # NMC
}


def pick_fill_and_font(prog, bachelor_alt, english_alt):
    """Decide row colour and font based on programme properties."""
    nmc    = prog.get('nmc', 'No')
    level  = prog.get('level', 'Bachelor')
    medium = prog.get('medium', 'English').lower()

    if nmc == 'Yes':
        return FILLS['nmc_yes'], FONTS['nmc'], bachelor_alt, english_alt

    if level in ('Diploma', 'Associate'):
        return FILLS['diploma'], FONTS['diploma'], bachelor_alt, english_alt

    if medium != 'english':
        fill = FILLS[f'english{(english_alt % 2) + 1}']
        return fill, FONTS['english'], bachelor_alt, english_alt + 1

    fill = FILLS[f'bachelor{(bachelor_alt % 2) + 1}']
    return fill, FONTS['standard'], bachelor_alt + 1, english_alt


def write_excel(programmes, university_name, output_dir='results'):
    """
    Build and save the Excel file.
    Returns the output file path.
    """
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()

    # ── Main programmes sheet
    ws = wb.active
    ws.title = university_name[:30]

    # ── Header row
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = FILLS['header']
        cell.font = FONTS['header']
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[1].height = 22

    # ── Freeze header
    ws.freeze_panes = "A2"

    # ── Data rows
    bachelor_alt = 0
    english_alt  = 0

    for row_num, prog in enumerate(programmes, 2):
        name    = prog.get('name', '')
        level   = prog.get('level', 'Bachelor')
        url     = prog.get('url', '')
        dur     = prog.get('duration', '')
        fee     = prog.get('fee', '')
        medium  = prog.get('medium', 'English')
        nmc     = prog.get('nmc', 'No')

        fill, font, bachelor_alt, english_alt = pick_fill_and_font(
            prog, bachelor_alt, english_alt
        )

        values = [name, level, url, dur, fee, medium, nmc]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.fill = fill
            cell.alignment = Alignment(vertical='center', wrap_text=False)

            # URL column — hyperlink formula
            if col == 3 and value and value.startswith('http'):
                safe = value.replace('"', '%22')
                cell.value = f'=HYPERLINK("{safe}","{safe}")'
                cell.font = FONTS['url']
            else:
                cell.font = font

        ws.row_dimensions[row_num].height = 18

    # ── Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # ── Column widths
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Summary sheet
    ws2 = wb.create_sheet("Summary")

    bachelors  = [p for p in programmes if p.get('level') == 'Bachelor']
    associates = [p for p in programmes if p.get('level') == 'Associate']
    diplomas   = [p for p in programmes if p.get('level') == 'Diploma']
    nmc_yes    = [p for p in programmes if p.get('nmc') == 'Yes']
    nmc_no     = [p for p in programmes if p.get('nmc') == 'No']

    summary_data = [
        ("University",        university_name),
        ("Total Programmes",  len(programmes)),
        ("Bachelor's",        len(bachelors)),
        ("Associate",         len(associates)),
        ("Diploma",           len(diplomas)),
        ("NMC: Yes",          len(nmc_yes)),
        ("NMC: No",           len(nmc_no)),
    ]

    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 35

    # Header
    ws2.cell(row=1, column=1, value="Field").fill  = FILLS['header']
    ws2.cell(row=1, column=1).font = FONTS['header']
    ws2.cell(row=1, column=2, value="Value").fill  = FILLS['header']
    ws2.cell(row=1, column=2).font = FONTS['header']

    for i, (field, value) in enumerate(summary_data, 2):
        ws2.cell(row=i, column=1, value=field).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=value)
        ws2.row_dimensions[i].height = 18

    # ── Save
    safe_name = university_name.lower().replace(' ', '-').replace('/', '-')
    filename = os.path.join(output_dir, f"{safe_name}-programmes.xlsx")
    wb.save(filename)
    print(f"  ✅ Saved: {filename}")
    return filename
